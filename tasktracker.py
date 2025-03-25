
import tkinter as tk
from tkinter import messagebox, ttk
import time
import threading
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

class PomodoroApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ABCD Pomodoro Task Manager")
        self.root.geometry("1000x750")
        self.root.resizable(False, False)

        self.excel_file = "tasks.xlsx"
        self.daily_capacity = 14
        self.tasks_df = self.load_tasks_from_excel(self.excel_file)
        self.tasks = self.tasks_df["Task Name"].tolist()
        self.completed_pomodoros = {task: 0 for task in self.tasks}

        self.selected_task = tk.StringVar()
        self.timer_label = tk.Label(root, text="Timer: 25:00", font=("Helvetica", 20))
        self.timer_label.pack(pady=10)

        self.task_dropdown = ttk.Combobox(root, textvariable=self.selected_task, values=self.tasks, width=70)
        self.task_dropdown.pack(pady=5)
        self.task_dropdown.set("Select a task from your XLS file")

        self.start_button = tk.Button(root, text="Start Pomodoro", command=self.start_pomodoro)
        self.start_button.pack(pady=5)

        self.urgent_label = tk.Label(root, text="🚨 Urgent Tasks", font=("Helvetica", 14, "bold"), fg="red")
        self.urgent_label.pack(pady=10)

        self.urgent_table = ttk.Treeview(root, columns=("Task", "Remaining", "Due Date"), show="headings", height=4)
        self.urgent_table.heading("Task", text="Task")
        self.urgent_table.heading("Remaining", text="Pomodoros Left")
        self.urgent_table.heading("Due Date", text="Due Date")
        self.urgent_table.column("Task", width=400)
        self.urgent_table.column("Remaining", width=150)
        self.urgent_table.column("Due Date", width=150)
        self.urgent_table.pack(pady=5)

        self.schedule_label = tk.Label(root, text="📅 Today's Task Schedule", font=("Helvetica", 14))
        self.schedule_label.pack(pady=10)

        self.schedule_table = ttk.Treeview(root, columns=("Task", "Effort", "Urgent"), show="headings", height=10)
        self.schedule_table.heading("Task", text="Task")
        self.schedule_table.heading("Effort", text="Effort (Pomodoros)")
        self.schedule_table.heading("Urgent", text="Urgent")
        self.schedule_table.column("Task", width=400)
        self.schedule_table.column("Effort", width=150)
        self.schedule_table.column("Urgent", width=100)
        self.schedule_table.pack(pady=5)

        self.populate_schedule()

        self.stop_event = threading.Event()
        self.timer_thread = None

    def load_tasks_from_excel(self, filename):
        try:
            df = pd.read_excel(filename, engine="openpyxl")
            df = df[df["Priority (A/B/C/D)"].isin(["A", "B", "C"])]
            df = df[df["Status"].str.lower() != "completed"]
            df = df.sort_values(by="Priority (A/B/C/D)", key=lambda col: col.map({"A": 1, "B": 2, "C": 3}))
            df["Effort Estimate"] = pd.to_numeric(df["Effort Estimate"], errors="coerce").fillna(1).astype(int)
            df["Progress (%)"] = pd.to_numeric(df["Progress (%)"], errors="coerce").fillna(0).astype(float)
            df["Due Date"] = pd.to_datetime(df["Due Date"], errors="coerce", dayfirst=True)
            return df
        except Exception as e:
            messagebox.showerror("File Error", f"Could not read file: {e}")
            return pd.DataFrame(columns=["Task Name", "Effort Estimate", "Progress (%)", "Due Date"])

    def populate_schedule(self):
        for item in self.schedule_table.get_children():
            self.schedule_table.delete(item)
        for item in self.urgent_table.get_children():
            self.urgent_table.delete(item)

        today = datetime.today().weekday()
        weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        today_name = weekdays[today] if today < 5 else "Monday"
        today_date = datetime.today().date()

        today_schedule = []
        urgent_tasks = []

        for _, row in self.tasks_df.iterrows():
            task = row["Task Name"]
            effort = row["Effort Estimate"]
            progress = row["Progress (%)"]
            due_raw = row["Due Date"]
            due = None
            if pd.notnull(due_raw):
                try:
                    due = pd.to_datetime(due_raw, dayfirst=True).date()
                except Exception as e:
                    print(f"[WARN] Could not parse due date for task '{task}': {due_raw}")
            
            completed = round((progress / 100) * effort)
            
            remaining = effort - completed

            days_left = max((due - today_date).days, 1) if due else 5
            max_available = days_left * self.daily_capacity
            is_urgent = remaining > max_available or (progress < 50 and days_left <= 3)

            if is_urgent:
                urgent_tasks.append((task, remaining, due))
                today_schedule.insert(0, (task, remaining, "YES"))
            else:
                today_schedule.append((task, remaining, ""))

        for task, rem, due in urgent_tasks:
            self.urgent_table.insert("", "end", values=(task, rem, due))

        for task, rem, urg in today_schedule:
            self.schedule_table.insert("", "end", values=(task, rem, urg))

    def start_pomodoro(self):
        task = self.selected_task.get()
        if not task or task == "Select a task from your XLS file":
            messagebox.showwarning("Input Error", "Please select a task.")
            return

        self.stop_event.clear()
        self.timer_thread = threading.Thread(target=self.run_timer, args=(25*60, task))
        self.timer_thread.start()

    def run_timer(self, duration, task):
        remaining = duration
        while remaining > 0 and not self.stop_event.is_set():
            mins, secs = divmod(remaining, 60)
            time_str = f"Timer: {mins:02}:{secs:02}"
            self.timer_label.config(text=time_str)
            time.sleep(1)
            remaining -= 1

        if not self.stop_event.is_set():
            self.timer_label.config(text="Time's up!")
            self.completed_pomodoros[task] += 1
            self.update_progress_in_excel(task)
            messagebox.showinfo("Pomodoro Complete", f"Finished: {task}")
            self.tasks_df = self.load_tasks_from_excel(self.excel_file)
            self.populate_schedule()

    def update_progress_in_excel(self, task_name):
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active

            for row in range(2, ws.max_row + 1):
                if ws[f"A{row}"].value == task_name:
                    effort = ws[f"I{row}"].value or 1
                    completed = self.completed_pomodoros[task_name]
                    progress = min(100, round((completed / effort) * 100))
                    ws[f"J{row}"] = progress
                    break

            wb.save(self.excel_file)
        except Exception as e:
            messagebox.showerror("Write Error", f"Could not update progress: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = PomodoroApp(root)
    root.mainloop()
